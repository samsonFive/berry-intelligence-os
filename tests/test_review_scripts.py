from __future__ import annotations

from openpyxl import load_workbook

from app import main

FAKE_SOURCE = {
    "id": "source-test",
    "type": "keyword",
    "label": "Fictional keyword",
    "value": "fictional term",
    "berry_ids": [],
    "enabled": True,
}


def _fake_auto_captured(record_id: str, *, source_name="Fictional Publisher", origin_domain="fictional-publisher.invalid") -> dict:
    return {
        "id": record_id,
        "record_type": "evidence",
        "status": "published",
        "source_type": "news_search",
        "title": f"Fictional headline for {record_id} - Fictional Publisher",
        "source_name": source_name,
        "source_id": "source-test",
        "source_url": "https://news.google.com/rss/articles/fictional",
        "origin_domain": origin_domain,
        "published_date": "2026-01-01",
        "captured_date": "2026-01-01",
        "summary": "A fictional summary used to test the review-export round trip.",
        "why_it_matters": "",
        "submitted_by": "source-monitor:Fictional keyword",
        "berry_ids": [],
        "geography_ids": [],
        "entity_ids": [],
        "fact_ids": [],
        "relationship_ids": [],
        "strategic_question_ids": [],
        "tags": ["Fictional keyword"],
        "auto_captured": True,
        "validated": False,
        "priority": {dim: {"level": "none", "rationale": ""} for dim in main.PRIORITY_DIMENSIONS},
    }


def _isolate(monkeypatch, tmp_path) -> None:
    monkeypatch.setattr(main, "INBOX_DIR", tmp_path / "inbox")
    monkeypatch.setattr(main, "DATA_DIR", tmp_path / "data")


def test_export_for_review_writes_only_unreviewed_auto_captured_rows(monkeypatch, tmp_path) -> None:
    _isolate(monkeypatch, tmp_path)
    main.save_sources([FAKE_SOURCE])
    main.save_evidence(_fake_auto_captured("ev-unreviewed-1"))
    validated = _fake_auto_captured("ev-already-validated")
    validated["validated"] = True
    main.save_evidence(validated)
    manual = _fake_auto_captured("ev-manual")
    manual["auto_captured"] = False
    main.save_evidence(manual)

    import scripts.export_for_review as export_for_review

    output_path = tmp_path / "review.xlsx"
    monkeypatch.setattr("sys.argv", ["export_for_review.py", str(output_path)])
    export_for_review.main()

    wb = load_workbook(output_path)
    ws = wb.active
    ids = [row[-1] for row in ws.iter_rows(min_row=2, values_only=True)]
    assert ids == ["ev-unreviewed-1"]


def test_export_for_review_uses_real_publisher_and_domain_not_google(monkeypatch, tmp_path) -> None:
    _isolate(monkeypatch, tmp_path)
    main.save_sources([FAKE_SOURCE])
    main.save_evidence(_fake_auto_captured("ev-unreviewed-1", source_name="FreshPlaza", origin_domain="freshplaza.com"))

    import scripts.export_for_review as export_for_review

    output_path = tmp_path / "review.xlsx"
    monkeypatch.setattr("sys.argv", ["export_for_review.py", str(output_path)])
    export_for_review.main()

    wb = load_workbook(output_path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    row = dict(zip(header, next(ws.iter_rows(min_row=2, values_only=True))))
    assert row["publisher"] == "FreshPlaza"
    assert row["domain"] == "freshplaza.com"
    assert row["domain"] != "news.google.com"


def test_apply_review_decisions_validate_purge_and_block(monkeypatch, tmp_path) -> None:
    _isolate(monkeypatch, tmp_path)
    main.save_sources([FAKE_SOURCE])
    for i in range(3):
        main.save_evidence(_fake_auto_captured(f"ev-{i}", origin_domain=f"domain{i}.invalid"))

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["decision", "title", "publisher", "domain", "keyword_source", "published_date", "summary", "source_url", "notes", "id"])
    ws.append(["validate", "t0", "p", "domain0.invalid", "k", "2026-01-01", "s", "u", "", "ev-0"])
    ws.append(["purge", "t1", "p", "domain1.invalid", "k", "2026-01-01", "s", "u", "", "ev-1"])
    ws.append(["purge+block", "t2", "p", "domain2.invalid", "k", "2026-01-01", "s", "u", "", "ev-2"])
    xlsx_path = tmp_path / "reviewed.xlsx"
    wb.save(xlsx_path)

    import scripts.apply_review_decisions as apply_review_decisions

    monkeypatch.setattr("sys.argv", ["apply_review_decisions.py", str(xlsx_path)])
    apply_review_decisions.main()

    assert (main.DATA_DIR / "evidence" / "ev-0.json").exists()
    assert main.all_evidence()[0]["validated"] in (True, False)  # sanity: still valid JSON
    ev0 = [r for r in main.all_evidence() if r["id"] == "ev-0"][0]
    assert ev0["validated"] is True
    assert not (main.DATA_DIR / "evidence" / "ev-1.json").exists()
    assert not (main.DATA_DIR / "evidence" / "ev-2.json").exists()
    assert "domain2.invalid" in main.load_blocked_domains()
    source = main.load_sources()[0]
    assert source["validated_count"] == 1
    assert source["purged_count"] == 2


def test_apply_review_decisions_leaves_blank_rows_untouched(monkeypatch, tmp_path) -> None:
    _isolate(monkeypatch, tmp_path)
    main.save_sources([FAKE_SOURCE])
    main.save_evidence(_fake_auto_captured("ev-0"))

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["decision", "title", "publisher", "domain", "keyword_source", "published_date", "summary", "source_url", "notes", "id"])
    ws.append(["", "t0", "p", "d.invalid", "k", "2026-01-01", "s", "u", "", "ev-0"])
    xlsx_path = tmp_path / "reviewed.xlsx"
    wb.save(xlsx_path)

    import scripts.apply_review_decisions as apply_review_decisions

    monkeypatch.setattr("sys.argv", ["apply_review_decisions.py", str(xlsx_path)])
    apply_review_decisions.main()

    assert (main.DATA_DIR / "evidence" / "ev-0.json").exists()
    ev0 = [r for r in main.all_evidence() if r["id"] == "ev-0"][0]
    assert ev0["validated"] is False
