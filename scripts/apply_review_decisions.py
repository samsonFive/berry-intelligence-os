"""Apply decisions from an edited review spreadsheet (see
export_for_review.py) back to the evidence dataset.

Usage:
    python scripts/apply_review_decisions.py path/to/reviewed.xlsx

For every row with a non-blank "decision":
    validate      -- mark the record validated, credit the source's tally
    purge         -- delete the record, debit the source's tally
    purge+block   -- delete the record, debit the tally, and add its real
                     publisher domain to the block list so future captures
                     from that outlet are skipped

Rows left blank are untouched. Rows whose id no longer exists (e.g. already
actioned through the web UI) are reported, not treated as an error.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook  # noqa: E402

import app.main as bios  # noqa: E402

VALID_DECISIONS = {"validate", "purge", "purge+block"}


def main() -> None:
    if len(sys.argv) < 2:
        print("Usage: python scripts/apply_review_decisions.py path/to/reviewed.xlsx")
        raise SystemExit(1)

    wb = load_workbook(sys.argv[1])
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    col = {name: idx for idx, name in enumerate(header)}

    counts = {"validate": 0, "purge": 0, "purge+block": 0}
    skipped_blank = 0
    not_found = []
    bad_decision = []
    newly_blocked: set[str] = set()

    # Batched rather than written through bump_source_tally()/
    # add_blocked_domain() per row -- those each do their own full
    # load+save of sources.json/blocked_domains.json, which is fine for a
    # single web-UI click but rewrites a growing file hundreds of times in
    # a row for a spreadsheet this size (observed: minutes instead of
    # seconds for ~1600 rows). One load, accumulate in memory, one save.
    sources = bios.load_sources()
    sources_by_id = {s["id"]: s for s in sources}
    blocked_domains = set(bios.load_blocked_domains())

    for row in ws.iter_rows(min_row=2, values_only=True):
        decision = (row[col["decision"]] or "").strip().lower()
        record_id = row[col["id"]]
        if not decision:
            skipped_blank += 1
            continue
        if decision not in VALID_DECISIONS:
            bad_decision.append((record_id, decision))
            continue

        path = bios.DATA_DIR / "evidence" / f"{record_id}.json"
        if not path.exists():
            not_found.append(record_id)
            continue

        record = json.loads(path.read_text(encoding="utf-8"))
        source = sources_by_id.get(record.get("source_id"))

        if decision == "validate":
            record["validated"] = True
            path.write_text(json.dumps(record, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
            if source is not None:
                source["validated_count"] = source.get("validated_count", 0) + 1
            counts["validate"] += 1
        else:
            if decision == "purge+block":
                domain = record.get("origin_domain") or bios.domain_of(record.get("source_url", ""))
                if domain and domain not in bios.UNBLOCKABLE_DOMAINS:
                    blocked_domains.add(domain)
                    newly_blocked.add(domain)
            if source is not None:
                source["purged_count"] = source.get("purged_count", 0) + 1
            path.unlink()
            counts[decision] += 1

    bios.save_sources(sources)
    if newly_blocked:
        bios.save_blocked_domains(sorted(blocked_domains))

    print(f"Validated: {counts['validate']}")
    print(f"Purged: {counts['purge']}")
    print(f"Purged & blocked: {counts['purge+block']}")
    if newly_blocked:
        print(f"Newly blocked domains: {sorted(newly_blocked)}")
    print(f"Left blank (untouched): {skipped_blank}")
    if not_found:
        print(f"Skipped, id no longer exists ({len(not_found)}): {not_found[:10]}{'...' if len(not_found) > 10 else ''}")
    if bad_decision:
        print(f"Skipped, unrecognized decision value ({len(bad_decision)}): {bad_decision[:10]}")


if __name__ == "__main__":
    main()
