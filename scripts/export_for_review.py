"""Export the auto-captured, unreviewed evidence backlog to an Excel sheet
for offline review, since reviewing 1500+ records one at a time through the
web UI's Validate/Purge buttons isn't practical.

Usage:
    python scripts/export_for_review.py [output_path.xlsx]

The default output path is review/review-backlog.xlsx. review/ is kept out of
Git by .gitignore (except the one committed snapshot), so the default run does
not leave an untracked artifact at the repository root that could be committed
by accident.

Fill in the "decision" column for any row you've made a call on and leave
the rest blank -- blank rows are left untouched when you run
apply_review_decisions.py against the file. Valid decisions:
    validate      -- keep it, mark as reviewed
    purge         -- delete it
    purge+block   -- delete it and stop capturing from that domain going forward
"""
from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from openpyxl.worksheet.datavalidation import DataValidation  # noqa: E402

from app.main import all_evidence, domain_of, load_sources  # noqa: E402

DECISIONS = ["validate", "purge", "purge+block"]

COLUMNS = [
    ("decision", 14),
    ("title", 60),
    ("publisher", 20),
    ("domain", 26),
    ("keyword_source", 22),
    ("published_date", 13),
    ("summary", 80),
    ("source_url", 50),
    ("notes", 30),
    ("id", 46),
]


def main() -> None:
    output_path = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "review" / "review-backlog.xlsx"

    sources_by_id = {s["id"]: s for s in load_sources()}
    backlog = [r for r in all_evidence() if r.get("auto_captured") and not r.get("validated")]
    backlog.sort(
        key=lambda r: (
            r.get("origin_domain") or domain_of(r.get("source_url", "")),
            sources_by_id.get(r.get("source_id"), {}).get("label", ""),
            r.get("title", ""),
        )
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Review backlog"

    header = [name for name, _ in COLUMNS]
    ws.append(header)
    for col_idx, (_, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for record in backlog:
        source = sources_by_id.get(record.get("source_id"), {})
        row = {
            "decision": "",
            "title": record.get("title", ""),
            "publisher": record.get("source_name", ""),
            "domain": record.get("origin_domain") or domain_of(record.get("source_url", "")),
            "keyword_source": source.get("label", ""),
            "published_date": record.get("published_date") or record.get("captured_date", ""),
            "summary": record.get("summary", ""),
            "source_url": record.get("source_url", ""),
            "notes": "",
            "id": record["id"],
        }
        ws.append([row[name] for name, _ in COLUMNS])

    # Only attach the decision-column dropdown when there is at least one data
    # row. On a clean/empty clone the backlog is empty (max_row == 1, header
    # only), and DataValidation.add("A2:A1") raises "1 must be greater than 2",
    # so the documented default run crashed instead of writing an empty sheet.
    if ws.max_row >= 2:
        decision_col_letter = get_column_letter(1)
        validation = DataValidation(type="list", formula1=f'"{",".join(DECISIONS)}"', allow_blank=True)
        ws.add_data_validation(validation)
        validation.add(f"{decision_col_letter}2:{decision_col_letter}{ws.max_row}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Wrote {len(backlog)} rows to {output_path}")


if __name__ == "__main__":
    main()
