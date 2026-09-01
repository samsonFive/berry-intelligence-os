"""UPOV PLUTO — operator-supplied Premium download only.

Do not scrape authenticated PLUTO pages. There is no public consumer API.
Premium (CHF 750/year) permits Excel download in the UI. Terms of use
(updated 2025-06-03) forbid creating derivative databases and commercial
dissemination except up to 100 records, and require written permission to
distribute analysis of more than 100 records.

This parser accepts an operator-exported workbook or JSON rows, hard-caps
at 100 records, and maps onto variety-candidate fields. National PVP
authorities remain more authoritative than PLUTO.
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any, BinaryIO

from openpyxl import load_workbook

from app.services.variety_universe.registry_import import import_registry_rows

MAX_DISTRIBUTABLE_RECORDS = 100
PREMIUM_CHF_PER_YEAR = 750
SOURCE_ID = "upov-pluto-operator-export"
SOURCE_LABEL = "UPOV PLUTO operator export"

LICENSING_FLAGS = (
    "no_public_api",
    "no_authenticated_scrape",
    "premium_download_ui_only",
    "derivative_database_not_authorized",
    "commercial_dissemination_not_authorized",
    "distribution_cap_100_records",
    "written_permission_required_over_100",
    "not_official_publication",
    "members_not_obliged_to_supply",
    "saas_resale_vendor_review_required",
)


class UpovPlutoError(RuntimeError):
    pass


def _header_map(values: list[Any]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    aliases = {
        "denomination": ("denomination", "variety denomination", "variety name"),
        "crop": ("crop", "botanical name", "species"),
        "applicant": ("applicant", "owner", "breeder"),
        "application_number": ("application number", "application no", "appl. number"),
        "grant_number": ("grant number", "title number", "certificate number"),
        "status": ("status", "current status"),
        "application_date": ("application date", "filing date"),
        "grant_date": ("grant date", "granting date"),
        "jurisdiction": ("jurisdiction", "country", "office"),
    }
    folded = [str(value or "").strip().lower() for value in values]
    for field, names in aliases.items():
        for index, cell in enumerate(folded):
            if cell in names:
                mapping[field] = index
                break
    return mapping


def parse_operator_export(data: bytes | BinaryIO | Path | list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Parse a Premium UI download. Hard-cap 100. Refuse HTML/login pages."""
    if isinstance(data, list):
        rows = [row for row in data if isinstance(row, dict)]
    else:
        if isinstance(data, Path):
            raw = data.read_bytes()
        elif isinstance(data, (bytes, bytearray)):
            raw = bytes(data)
        else:
            raw = data.read()
        if raw.lstrip().startswith(b"<") or b"WIPO" in raw[:200] and b"login" in raw.lower():
            raise UpovPlutoError("refusing HTML/login payload; do not scrape PLUTO")
        workbook = load_workbook(BytesIO(raw), data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        header = [sheet.cell(1, col).value for col in range(1, (sheet.max_column or 1) + 1)]
        columns = _header_map(header)
        if "denomination" not in columns:
            raise UpovPlutoError("operator export is missing a denomination/variety column")
        rows = []
        for index in range(2, (sheet.max_row or 1) + 1):
            values = [sheet.cell(index, col).value for col in range(1, (sheet.max_column or 1) + 1)]
            get = lambda key: values[columns[key]] if key in columns else None
            name = str(get("denomination") or "").strip()
            if not name:
                continue
            rows.append(
                {
                    "denomination": name,
                    "crop": str(get("crop") or "").strip(),
                    "applicant": str(get("applicant") or "").strip(),
                    "application_number": str(get("application_number") or "").strip(),
                    "grant_number": str(get("grant_number") or "").strip(),
                    "status": str(get("status") or "").strip(),
                    "application_date": str(get("application_date") or "").strip(),
                    "grant_date": str(get("grant_date") or "").strip(),
                    "jurisdiction": str(get("jurisdiction") or "").strip(),
                }
            )
    if len(rows) > MAX_DISTRIBUTABLE_RECORDS:
        raise UpovPlutoError(
            f"refusing {len(rows)} PLUTO rows; Premium ToS cap is {MAX_DISTRIBUTABLE_RECORDS} "
            "without written UPOV permission"
        )
    mapped: list[dict[str, Any]] = []
    for row in rows:
        name = str(row.get("denomination") or row.get("candidate_name") or "").strip()
        mapped.append(
            {
                "candidate_name": name,
                "denomination": name,
                "berry_id": row.get("berry_id") or "",
                "applicant": row.get("applicant") or "",
                "breeder_owner": row.get("applicant") or row.get("breeder_owner") or "",
                "application_number": row.get("application_number") or "",
                "grant_number": row.get("grant_number") or "",
                "application_date": row.get("application_date") or "",
                "grant_date": row.get("grant_date") or "",
                "registration_status": row.get("status") or row.get("registration_status") or "",
                "jurisdiction": row.get("jurisdiction") or "",
                "source_id": SOURCE_ID,
                "source_label": SOURCE_LABEL,
                "source_url": "https://www.upov.int/en/find-and-explore/databases/pluto-search",
                "source_type": "plant_breeders_rights_record",
                "source_tier": "tier_1_registry",
                "notes": "PLUTO is a cross-jurisdiction index, not the official national publication.",
            }
        )
    return mapped


def import_operator_rows(
    rows: list[dict[str, Any]],
    *,
    varieties: list[dict[str, Any]],
    inbox_dir: Path,
) -> dict[str, Any]:
    if len(rows) > MAX_DISTRIBUTABLE_RECORDS:
        raise UpovPlutoError("distribution cap exceeded")
    return import_registry_rows(rows, varieties=varieties, inbox_dir=inbox_dir)
