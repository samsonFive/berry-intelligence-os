"""USDA Plant Variety Protection Office — structured monthly status report.

There is no documented public PVPO API (verified 2026-09-01 against AMS
PVPO pages). Public structured access is the monthly Application Status
Report XLSX. ePVP is applicant-only. CMS/POD are search UIs, not APIs.
This module parses the XLSX and maps berry rows onto variety-candidate
fields. It never writes trusted Evidence or onboarded Sources.
"""

from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path
from typing import Any, BinaryIO, Callable

from openpyxl import load_workbook

from app.services.variety_universe.registry_import import build_candidate, import_registry_rows

STATUS_REPORT_URL = "https://www.ams.usda.gov/sites/default/files/media/PVPOApplicationStatus.xlsx"
SOURCE_ID = "usda-pvpo-application-status"
SOURCE_LABEL = "USDA PVPO Application Status Report"
JURISDICTION = "US"
NO_PUBLIC_API = True
UPDATE_CADENCE = "monthly"

# Header is row 2; column A is empty. Documented by live 2026-09-01 file.
HEADER_ROW = 2
COL = {
    "application_number": 2,
    "variety_name": 3,
    "experimental_name": 4,
    "scientific_name": 5,
    "common_name": 6,
    "applicant": 7,
    "application_date": 8,
    "certificate_status": 10,
    "status_date": 11,
    "issued_date": 12,
}

BERRY_COMMON = {
    "blueberry": "berry-blueberry",
    "blueberry, rootstock": "berry-blueberry",
    "blueberry, highbush": "berry-blueberry",
    "blueberry, rabbiteye": "berry-blueberry",
    "strawberry": "berry-strawberry",
    "raspberry": "berry-raspberry",
    "raspberry, red": "berry-raspberry",
    "raspberry, black": "berry-raspberry",
    "blackberry": "berry-blackberry",
}
BERRY_PREFIXES = (
    ("blueberry", "berry-blueberry"),
    ("strawberry", "berry-strawberry"),
    ("raspberry", "berry-raspberry"),
    ("blackberry", "berry-blackberry"),
)
SCI_HINTS = (
    ("vaccinium", "berry-blueberry"),
    ("fragaria", "berry-strawberry"),
    ("rubus idaeus", "berry-raspberry"),
    ("rubus occidentalis", "berry-raspberry"),
    ("rubus ursinus", "berry-blackberry"),
    ("rubus allegheniensis", "berry-blackberry"),
)


class UsdaPvpoError(RuntimeError):
    pass


def berry_id_for(*, common_name: str, scientific_name: str) -> str | None:
    common = (common_name or "").strip().lower()
    if common in BERRY_COMMON:
        return BERRY_COMMON[common]
    for prefix, berry_id in BERRY_PREFIXES:
        if common == prefix or common.startswith(prefix + ",") or common.startswith(prefix + " "):
            return berry_id
    sci = (scientific_name or "").strip().lower()
    for hint, berry_id in SCI_HINTS:
        if hint in sci:
            return berry_id
    return None


def _cell(ws: Any, row: int, key: str) -> Any:
    return ws.cell(row, COL[key]).value


def parse_status_workbook(data: bytes | BinaryIO | Path) -> list[dict[str, Any]]:
    """Parse the monthly AMS XLSX. Berry rows only. No HTML scrape."""
    if isinstance(data, Path):
        workbook = load_workbook(data, data_only=True)
    elif isinstance(data, (bytes, bytearray)):
        workbook = load_workbook(BytesIO(data), data_only=True)
    else:
        workbook = load_workbook(data, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header = [sheet.cell(HEADER_ROW, col).value for col in range(1, 14)]
    if "Application #" not in header and "Variety Name" not in header:
        raise UsdaPvpoError("PVPO status workbook is missing the Application # / Variety Name header")
    rows: list[dict[str, Any]] = []
    for index in range(HEADER_ROW + 1, (sheet.max_row or 0) + 1):
        common = str(_cell(sheet, index, "common_name") or "").strip()
        scientific = str(_cell(sheet, index, "scientific_name") or "").strip()
        berry_id = berry_id_for(common_name=common, scientific_name=scientific)
        if not berry_id:
            continue
        application_number = str(_cell(sheet, index, "application_number") or "").strip()
        name = str(_cell(sheet, index, "variety_name") or "").strip()
        if not application_number or not name:
            continue
        issued = _cell(sheet, index, "issued_date")
        rows.append(
            {
                "candidate_name": name,
                "denomination": name,
                "trade_name": str(_cell(sheet, index, "experimental_name") or "").strip(),
                "berry_id": berry_id,
                "crop": common,
                "scientific_name": scientific,
                "applicant": str(_cell(sheet, index, "applicant") or "").strip(),
                "breeder_owner": str(_cell(sheet, index, "applicant") or "").strip(),
                "application_number": application_number,
                "grant_number": application_number if issued else "",
                "application_date": str(_cell(sheet, index, "application_date") or ""),
                "grant_date": str(issued or ""),
                "status_date": str(_cell(sheet, index, "status_date") or ""),
                "registration_status": str(_cell(sheet, index, "certificate_status") or "").strip(),
                "jurisdiction": JURISDICTION,
                "geography_id": "geography-united-states",
                "source_id": SOURCE_ID,
                "source_label": SOURCE_LABEL,
                "source_url": STATUS_REPORT_URL,
                "source_type": "plant_breeders_rights_record",
                "source_tier": "tier_1_national_register",
            }
        )
    return rows


def fetch_status_report(*, get: Callable[..., Any] | None = None, timeout: float = 45.0) -> bytes:
    import httpx

    fetcher = get or httpx.get
    response = fetcher(
        STATUS_REPORT_URL,
        timeout=timeout,
        headers={"User-Agent": "berry-intelligence-os-pvpo/0.1"},
        follow_redirects=True,
    )
    response.raise_for_status()
    return response.content


def import_berry_rows(
    rows: list[dict[str, Any]],
    *,
    varieties: list[dict[str, Any]],
    inbox_dir: Path,
) -> dict[str, Any]:
    """Inbox Variety candidates only. Never trusted Evidence."""
    return import_registry_rows(rows, varieties=varieties, inbox_dir=inbox_dir)


def _load_variety_entities(data_dir: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    root = data_dir / "entities"
    if not root.is_dir():
        return rows
    for path in root.rglob("*.json"):
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        if isinstance(payload, dict) and payload.get("entity_type") in {None, "variety"} and payload.get("id"):
            if payload.get("entity_type") == "variety" or "variety" in str(payload.get("id")):
                rows.append(payload)
    return [row for row in rows if row.get("entity_type") == "variety"]


def _parse_date(value: Any) -> str:
    text = str(value or "").strip()
    return text[:10] if text else ""


def summarize_berry_import(
    rows: list[dict[str, Any]],
    import_report: dict[str, Any],
) -> dict[str, Any]:
    from app.services.authoritative_registries.events import classify_pvp_event

    names = sorted({str(row.get("denomination") or row.get("candidate_name") or "") for row in rows if row.get("candidate_name")})
    filings = [_parse_date(row.get("application_date")) for row in rows]
    updates = [_parse_date(row.get("status_date") or row.get("grant_date")) for row in rows]
    events = [classify_pvp_event(row) for row in rows]
    candidates = import_report.get("candidates") or []
    matched = [
        row
        for row in candidates
        if row.get("candidate_canonical_match") or row.get("identity_state") == "possible_alias"
    ]
    ambiguous = [
        row
        for row in candidates
        if row.get("identity_state") in {"possible_alias", "unknown"}
    ]
    return {
        "state": "ok",
        "source_url": STATUS_REPORT_URL,
        "layer": "AUTHORITATIVE_REGISTRY",
        "raw_berry_records": len(rows),
        "distinct_variety_names": len(names),
        "variety_names": names[:80],
        "matched_canonical": len(matched),
        "candidates": int(import_report.get("written_count") or 0),
        "built_count": int(import_report.get("built_count") or 0),
        "ambiguous_identity": len(ambiguous),
        "distinct_new": int(import_report.get("distinct_new") or 0),
        "possible_alias": int(import_report.get("possible_alias") or 0),
        "unknown": int(import_report.get("unknown") or 0),
        "newest_filing": max((item for item in filings if item), default=None),
        "newest_update": max((item for item in updates if item), default=None),
        "event_counts": {
            kind: sum(1 for event in events if event["event_kind"] == kind)
            for kind in {event["event_kind"] for event in events}
        },
        "auto_confirmed": False,
        "trust_state": "UNREVIEWED_REGISTRY",
    }


def run_bounded_import(
    *,
    data_dir: Path,
    inbox_dir: Path,
    persist: bool = True,
    get: Callable[..., Any] | None = None,
) -> dict[str, Any]:
    """Fetch the official XLSX, map berry rows, propose candidates only."""
    raw = fetch_status_report(get=get)
    rows = parse_status_workbook(raw)
    varieties = _load_variety_entities(data_dir)
    if persist:
        imported = import_berry_rows(rows, varieties=varieties, inbox_dir=inbox_dir)
    else:
        built = [build_candidate(row, varieties=varieties) for row in rows]
        imported = {
            "input_count": len(rows),
            "built_count": len(built),
            "written_count": 0,
            "written_ids": [],
            "rejected_count": sum(1 for row in built if row.get("status") == "rejected"),
            "distinct_new": sum(1 for row in built if row.get("identity_state") == "distinct"),
            "possible_alias": sum(1 for row in built if row.get("identity_state") == "possible_alias"),
            "unknown": sum(1 for row in built if row.get("identity_state") == "unknown"),
            "candidates": built,
            "state": "dry_run",
        }
    summary = summarize_berry_import(rows, imported)
    if not persist:
        summary["state"] = "dry_run"
        summary["candidates"] = 0
    summary["bytes_downloaded"] = len(raw)
    summary["canonical_varieties_checked"] = len(varieties)
    return summary
